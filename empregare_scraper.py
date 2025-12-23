from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
import pandas as pd

cargo = input("Digite o cargo desejado para busca de vagas: ")
localidade = input("Digite a localidade desejada para busca de vagas (cidade, estado): ")
rolagem = int(input("Quantas vezes deseja rolar a página para carregar mais vagas? (+10 vagas por rolagem): "))

driver = webdriver.Chrome()
driver.get(f'https://www.empregare.com/pt-br/vagas?query={cargo}&localidade={localidade}')
time.sleep(3)

print(f"https://www.empregare.com/pt-br/vagas?query={cargo}&localidade={localidade}")

for _ in ([] if rolagem == 0 else range(rolagem)):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

soup = BeautifulSoup(driver.page_source, 'html.parser')

# já foi usado ent acho que não precisa mais...
driver.quit()

vagas = soup.find_all('a', class_='text-decoration-none')

# Criação de lista para armazenar as vagas e posteriormente criar um xslx
Dicionario_vagas = []
quantidade_colunas = 7
quantidade_linhas = 0

for num, vaga in enumerate(vagas, 1):
    
    titulo = vaga.find(class_='fw-bold fs-4 titulo-vaga text-truncate mb-0 me-3').get_text(strip=True)
    empresa = vaga.find(class_='text-truncate card-vaga-empresa').get_text(strip=True)
    data_publicacao = vaga.find(class_='texto-data-card').get_text(strip=True)
    local = vaga.find(class_='card-cidades').get_text(strip=True)
    modalidade_elem = vaga.find(class_='badge bg-span-card-vaga rounded-5 px-2 me-4')
    modalidade = modalidade_elem.get_text(strip=True) if modalidade_elem else "Não informado"
    salario_elem = vaga.find(class_='d-none d-md-flex badge bg-span-card-vaga rounded-5 px-2')
    salario = salario_elem.get_text(strip=True) if salario_elem else "Não informado"
    link = vaga['href']
    
    Dicionario_vagas.append({
        'Título': titulo,
        'Empresa': empresa,
        'Data de Publicação': data_publicacao,
        'Local': local,
        'Modalidade': modalidade,
        'Salário': salario,
        'Link': f"https://www.empregare.com{link}"
    })
    
    pass
quantidade_linhas = num

print(f'Foram encontradas {quantidade_linhas} vagas para o cargo de {cargo} em {localidade}.')

dataframe = pd.DataFrame(Dicionario_vagas)
dataframe.to_excel('Vagas_Empregare.xlsx', index=False, sheet_name='Vagas')

workbook = load_workbook('Vagas_Empregare.xlsx')
worksheet = workbook['Vagas']

for cell in worksheet[1]:  # Primeira linha
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for row in range(2, quantidade_linhas + 2):  # Linha 2 até última (precisa ser +2 por causa do cabeçalho)
    cell = worksheet[f'G{row}']
    worksheet[f'G{row}'].hyperlink = Dicionario_vagas[row - 2]['Link']
    cell.value = '🔗 Abrir Vaga'
    cell.style = 'Hyperlink'

LARGURAS_FIXAS = {
    'A': 42, 'B': 42, 'C': 22, 'D': 14, 
    'E': 16, 'F': 24, 'G': 15
}
for col, width in LARGURAS_FIXAS.items():
    worksheet.column_dimensions[col].width = width

tab = Table(displayName="TabelaVagas", ref=f"A1:G{quantidade_linhas}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
worksheet.add_table(tab)

workbook.save('Vagas_Empregare.xlsx')